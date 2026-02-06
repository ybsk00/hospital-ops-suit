"""
외래예약 엑셀 파서
EMR에서 내보낸 외래예약 엑셀 파일을 파싱하여 dict 리스트로 변환한다.
"""
import logging
from datetime import datetime

from openpyxl import load_workbook

logger = logging.getLogger("parser.outpatient")

# ──────────────────────────────────────────
# 헤더 매핑 (EMR 컬럼명 → 내부 필드명)
# ──────────────────────────────────────────
HEADER_MAP: dict[str, str] = {
    "환자번호": "emrPatientId",
    "환자ID": "emrPatientId",
    "EMR_ID": "emrPatientId",
    "환자명": "patientName",
    "이름": "patientName",
    "예약일": "appointmentDate",
    "예약일자": "appointmentDate",
    "진료일": "appointmentDate",
    "시작시간": "startTime",
    "예약시간": "startTime",
    "시작": "startTime",
    "종료시간": "endTime",
    "종료": "endTime",
    "담당의": "doctorName",
    "담당의사": "doctorName",
    "의사": "doctorName",
    "의사ID": "emrDoctorId",
    "EMR의사ID": "emrDoctorId",
    "진료실": "clinicRoomName",
    "진료실명": "clinicRoomName",
    "예약상태": "status",
    "상태": "status",
    "비고": "notes",
    "메모": "notes",
    "EMR예약ID": "emrAppointmentId",
    "예약번호": "emrAppointmentId",
}

REQUIRED_FIELDS = {"emrPatientId", "patientName", "appointmentDate", "startTime"}


def _detect_header_row(ws, max_scan: int = 10) -> tuple[int, dict[str, int]]:
    """첫 10행 내에서 헤더 행을 자동 감지한다."""
    for row_idx in range(1, max_scan + 1):
        cells = [str(cell.value or "").strip() for cell in ws[row_idx]]
        mapping: dict[str, int] = {}
        for col_idx, cell_val in enumerate(cells):
            if cell_val in HEADER_MAP:
                field = HEADER_MAP[cell_val]
                if field not in mapping:
                    mapping[field] = col_idx
        if len(mapping) >= 3:
            return row_idx, mapping
    raise ValueError("헤더 행을 찾을 수 없습니다. EMR 외래예약 엑셀 형식을 확인하세요.")


def _normalize_date(value) -> str | None:
    """날짜 값을 YYYY-MM-DD 문자열로 변환한다."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            continue
    return None


def _normalize_time(value) -> str | None:
    """시간 값을 HH:MM 문자열로 변환한다."""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.strftime("%H:%M")
    s = str(value).strip()
    # HH:MM:SS or HH:MM
    for fmt in ("%H:%M:%S", "%H:%M", "%H%M"):
        try:
            return datetime.strptime(s, fmt).strftime("%H:%M")
        except ValueError:
            continue
    return None


def _normalize_status(value) -> str:
    """EMR 상태를 시스템 상태로 매핑한다."""
    if not value:
        return "BOOKED"
    s = str(value).strip().upper()
    status_map = {
        "예약": "BOOKED",
        "BOOKED": "BOOKED",
        "접수": "CHECKED_IN",
        "CHECKED_IN": "CHECKED_IN",
        "완료": "COMPLETED",
        "COMPLETED": "COMPLETED",
        "취소": "CANCELLED",
        "CANCELLED": "CANCELLED",
        "미방문": "NO_SHOW",
        "NO_SHOW": "NO_SHOW",
        "변경": "CHANGED",
        "CHANGED": "CHANGED",
    }
    return status_map.get(s, "BOOKED")


def parse_outpatient_file(file_path: str) -> list[dict]:
    """외래예약 엑셀 파일을 파싱하여 dict 리스트를 반환한다."""
    logger.info(f"외래예약 파싱 시작: {file_path}")
    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    try:
        header_row, col_map = _detect_header_row(ws)
    except ValueError as e:
        logger.error(str(e))
        return []

    logger.info(f"헤더 감지 완료 (행 {header_row}): {list(col_map.keys())}")

    rows = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=header_row + 1), start=header_row + 1):
        cells = [cell.value for cell in row]

        # 빈 행 건너뛰기
        if all(c is None for c in cells):
            continue

        try:
            record: dict = {"_row": row_idx}

            for field, col_idx in col_map.items():
                val = cells[col_idx] if col_idx < len(cells) else None
                record[field] = val

            # 필수 필드 확인
            emr_id = str(record.get("emrPatientId", "") or "").strip()
            if not emr_id:
                record["_error"] = "환자번호 누락"
                rows.append(record)
                continue

            record["emrPatientId"] = emr_id
            record["patientName"] = str(record.get("patientName", "") or "").strip()

            # 날짜 정규화
            apt_date = _normalize_date(record.get("appointmentDate"))
            if not apt_date:
                record["_error"] = "예약일 형식 오류"
                rows.append(record)
                continue
            record["appointmentDate"] = apt_date

            # 시간 정규화
            start_time = _normalize_time(record.get("startTime"))
            if not start_time:
                record["_error"] = "시작시간 형식 오류"
                rows.append(record)
                continue
            record["startTime"] = start_time

            end_time = _normalize_time(record.get("endTime"))
            if not end_time:
                # 기본 30분 진료
                h, m = map(int, start_time.split(":"))
                m += 30
                if m >= 60:
                    h += 1
                    m -= 60
                end_time = f"{h:02d}:{m:02d}"
            record["endTime"] = end_time

            # 상태 정규화
            record["status"] = _normalize_status(record.get("status"))

            # 기타 필드
            record["doctorName"] = str(record.get("doctorName", "") or "").strip()
            record["emrDoctorId"] = str(record.get("emrDoctorId", "") or "").strip() or None
            record["clinicRoomName"] = str(record.get("clinicRoomName", "") or "").strip() or None
            record["notes"] = str(record.get("notes", "") or "").strip() or None
            record["emrAppointmentId"] = str(record.get("emrAppointmentId", "") or "").strip() or None

            rows.append(record)

        except Exception as e:
            logger.warning(f"행 {row_idx} 파싱 실패: {e}")
            rows.append({"_row": row_idx, "_error": f"파싱 오류: {str(e)}"})

    wb.close()
    logger.info(f"외래예약 파싱 완료: {len(rows)}행")
    return rows
