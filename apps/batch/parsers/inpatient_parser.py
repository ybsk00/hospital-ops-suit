"""
입원현황 엑셀 파서
EMR에서 내보낸 입원환자 목록 XLSX 파일을 파싱한다.
"""
import logging
from datetime import datetime
from typing import Any

import openpyxl

logger = logging.getLogger(__name__)

# EMR 엑셀 헤더 → 내부 필드 매핑
HEADER_MAP: dict[str, str] = {
    "환자번호": "emrPatientId",
    "환자ID": "emrPatientId",
    "EMR_ID": "emrPatientId",
    "환자명": "name",
    "이름": "name",
    "생년월일": "dob",
    "성별": "sex",
    "연락처": "phone",
    "전화번호": "phone",
    "입원일": "admitDate",
    "입원일자": "admitDate",
    "퇴원예정일": "plannedDischargeDate",
    "담당의": "attendingDoctor",
    "담당의사": "attendingDoctor",
    "병동": "wardName",
    "호실": "roomName",
    "베드": "bedLabel",
    "침상": "bedLabel",
    "상태": "status",
    "비고": "notes",
}

REQUIRED_FIELDS = {"emrPatientId", "name", "dob", "sex", "admitDate"}


def detect_header_row(ws) -> tuple[int, dict[str, int]]:
    """
    워크시트에서 헤더 행을 자동 감지한다.
    처음 10행 내에서 HEADER_MAP의 키와 3개 이상 매칭되는 행을 헤더로 판단.
    Returns: (헤더_행_번호, {내부필드명: 열_인덱스})
    """
    for row_idx in range(1, min(11, ws.max_row + 1)):
        col_map: dict[str, int] = {}
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is None:
                continue
            header_text = str(cell_value).strip()
            if header_text in HEADER_MAP:
                col_map[HEADER_MAP[header_text]] = col_idx

        if len(col_map) >= 3:
            return row_idx, col_map

    raise ValueError("헤더 행을 찾을 수 없습니다. EMR 엑셀 파일 형식을 확인하세요.")


def parse_cell_date(value: Any) -> datetime | None:
    """셀 값을 datetime으로 변환"""
    if value is None:
        return None
    if isinstance(value, datetime):
        return value
    text = str(value).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def parse_row(row_values: dict[str, Any], row_number: int) -> dict[str, Any]:
    """
    한 행을 파싱하여 정규화된 딕셔너리를 반환한다.
    유효성 검사 실패 시 errors 리스트를 포함.
    """
    errors: list[str] = []
    result: dict[str, Any] = {"_rowNumber": row_number, "_errors": errors}

    # emrPatientId
    emr_id = row_values.get("emrPatientId")
    if emr_id is None or str(emr_id).strip() == "":
        errors.append("환자번호가 비어있습니다.")
    else:
        result["emrPatientId"] = str(emr_id).strip()

    # name
    name = row_values.get("name")
    if name is None or str(name).strip() == "":
        errors.append("환자명이 비어있습니다.")
    else:
        result["name"] = str(name).strip()

    # dob
    dob = parse_cell_date(row_values.get("dob"))
    if dob is None:
        errors.append("생년월일이 올바르지 않습니다.")
    else:
        result["dob"] = dob

    # sex
    sex = row_values.get("sex")
    if sex is None or str(sex).strip() == "":
        errors.append("성별이 비어있습니다.")
    else:
        sex_str = str(sex).strip()
        if sex_str in ("남", "M", "male", "남자"):
            result["sex"] = "M"
        elif sex_str in ("여", "F", "female", "여자"):
            result["sex"] = "F"
        else:
            result["sex"] = sex_str

    # phone (optional)
    phone = row_values.get("phone")
    if phone is not None and str(phone).strip():
        result["phone"] = str(phone).strip().replace("-", "")

    # admitDate
    admit_date = parse_cell_date(row_values.get("admitDate"))
    if admit_date is None:
        errors.append("입원일이 올바르지 않습니다.")
    else:
        result["admitDate"] = admit_date

    # plannedDischargeDate (optional)
    planned = parse_cell_date(row_values.get("plannedDischargeDate"))
    if planned:
        result["plannedDischargeDate"] = planned

    # attendingDoctor (optional)
    doctor = row_values.get("attendingDoctor")
    if doctor and str(doctor).strip():
        result["attendingDoctor"] = str(doctor).strip()

    # ward/room/bed (optional)
    for field in ("wardName", "roomName", "bedLabel"):
        val = row_values.get(field)
        if val and str(val).strip():
            result[field] = str(val).strip()

    # notes (optional)
    notes = row_values.get("notes")
    if notes and str(notes).strip():
        result["notes"] = str(notes).strip()

    return result


def parse_inpatient_file(file_path: str) -> list[dict[str, Any]]:
    """
    입원현황 엑셀 파일을 파싱하여 행 데이터 리스트를 반환한다.
    """
    logger.info(f"파싱 시작: {file_path}")
    wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active

    if ws is None:
        raise ValueError("워크시트를 찾을 수 없습니다.")

    header_row, col_map = detect_header_row(ws)
    logger.info(f"헤더 행: {header_row}, 매핑: {col_map}")

    # 필수 필드 검증
    missing = REQUIRED_FIELDS - set(col_map.keys())
    if missing:
        raise ValueError(f"필수 컬럼이 누락되었습니다: {missing}")

    rows: list[dict[str, Any]] = []
    for row_idx in range(header_row + 1, ws.max_row + 1):
        # 빈 행 건너뛰기
        first_val = ws.cell(row=row_idx, column=list(col_map.values())[0]).value
        if first_val is None:
            continue

        row_values: dict[str, Any] = {}
        for field_name, col_idx in col_map.items():
            row_values[field_name] = ws.cell(row=row_idx, column=col_idx).value

        parsed = parse_row(row_values, row_idx)
        rows.append(parsed)

    wb.close()
    logger.info(f"파싱 완료: {len(rows)}건")
    return rows
