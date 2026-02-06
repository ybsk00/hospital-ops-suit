"""
파일 유효성 검증
- 파일 수신 완료 확인 (done_signal / stable_size)
- XLSX 무결성 검사
- SHA-256 중복 체크
"""
import hashlib
import logging
import os
import time

import openpyxl

from config import FILE_STABLE_WAIT_SEC, RECEIPT_MODE

logger = logging.getLogger(__name__)


def compute_sha256(file_path: str) -> str:
    """파일의 SHA-256 해시를 계산한다."""
    sha = hashlib.sha256()
    with open(file_path, "rb") as f:
        while chunk := f.read(8192):
            sha.update(chunk)
    return sha.hexdigest()


def is_file_ready(file_path: str) -> bool:
    """
    파일이 수신 완료 상태인지 확인한다.
    - done_signal: .done 시그널 파일 존재 여부
    - stable_size: 파일 크기가 일정 시간 동안 변하지 않으면 완료
    """
    if RECEIPT_MODE == "done_signal":
        done_path = file_path + ".done"
        return os.path.exists(done_path)
    elif RECEIPT_MODE == "stable_size":
        try:
            size1 = os.path.getsize(file_path)
            time.sleep(FILE_STABLE_WAIT_SEC)
            size2 = os.path.getsize(file_path)
            return size1 == size2 and size1 > 0
        except OSError:
            return False
    else:
        # eof_marker 또는 기타: 파일 존재 + 크기 > 0
        return os.path.exists(file_path) and os.path.getsize(file_path) > 0


def validate_xlsx(file_path: str) -> tuple[bool, str]:
    """
    XLSX 파일 무결성 검사.
    Returns: (성공여부, 에러메시지)
    """
    if not os.path.exists(file_path):
        return False, f"파일이 존재하지 않습니다: {file_path}"

    if not file_path.lower().endswith((".xlsx", ".xls")):
        return False, f"지원하지 않는 파일 형식입니다: {file_path}"

    try:
        wb = openpyxl.load_workbook(file_path, read_only=True)
        ws = wb.active
        if ws is None or ws.max_row is None or ws.max_row < 2:
            wb.close()
            return False, "데이터가 없는 빈 파일입니다."
        wb.close()
        return True, ""
    except Exception as e:
        return False, f"XLSX 파일 열기 실패: {str(e)}"


def check_duplicate(file_hash: str, conn) -> bool:
    """
    Import 테이블에서 동일 해시의 파일이 이미 처리되었는지 확인한다.
    Returns: True = 중복
    """
    with conn.cursor() as cur:
        cur.execute(
            'SELECT COUNT(*) FROM "Import" WHERE "fileHash" = %s AND "status" IN (%s, %s)',
            (file_hash, "SUCCESS", "PROCESSING"),
        )
        count = cur.fetchone()[0]
    return count > 0
